import openpyxl

# 새로운 엑셀 문서 만들기

# 워크북(Workbook) 객체 만들기. Workbook() 클래스를 사용해서 빈 워크북 객체 생성
workbook = openpyxl.Workbook()  # 워크북 객체 생성
print(workbook.sheetnames)  # ['Sheet']. 워크북의 모든 sheet 출력. 새로 생성해서 하나의 sheet만 있음
# 주의할 점은 엑셀로 새문서를 만들면 sheet 이름이 'Sheet1' 이지만
# openpyxl 이용해서 만들면 sheet 이름이 'Sheet'

# 워크북의 변경내용을 새로운 파일에 저장
workbook.save('./output/create_workbook.xlsx')
