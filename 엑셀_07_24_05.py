import openpyxl
import openpyxl.utils.cell as ut_cell
from openpyxl.styles import Alignment, Font # 서식 지정을 위해 추가

# 새로운 엑셀 문서 만들기
# 1) 빈 워크북 객체 생성 2) 시트 객체 생성 3) 서식 지정해서 셀 데이터 입력 4) 엑셀 파일로 저장

# 1) 빈 워크북 객체 생성
workbook = openpyxl.Workbook()

# 2) 시트 객체 생성
# 워크북객체.create_sheet() : 시트 객체 생성. index : 시트 순서 값, title : 시트 이름 값
# 워크북객체.remove() : 시트 삭제

worksheet1 = workbook.create_sheet(index=0, title='Column') # 시트 객체 생성
print(workbook.sheetnames)  # ['Column', 'Sheet']. 워크북의 모든 sheet 출력. Sheet는 기본 생성된 sheet
workbook.remove(workbook['Sheet']) # Sheet 시트 삭제
print(workbook.sheetnames)  # ['Column']. 워크북의 모든 sheet 출력
worksheet2 = workbook.create_sheet(index=1, title='Row') # 시트 생성
print(workbook.sheetnames)  # ['Column', 'Row']. 워크북의 모든 sheet 출력
# 작업 내용을 정리하면 기본생성된 시트를 삭제하고, 'Column', 'Row' 시트 생성.

# 3) 서식 지정해서 셀 데이터 입력
# 기존 파일의 셀 데이터를 읽어올 때는 범위 지정이 필요 없지만, 빈 셀에 데이터 입력시에는 셀 범위를 지정
# min_row = 시작행, min_col = 시작열, max_row = 끝행, max_col = 끝열
for col in worksheet1.iter_cols(min_row=1, min_col=1, max_row=3, max_col=6) :
    # print(col)
    for each_cell in col :
        # 각 셀의 열 값을 문자로 변환해서 셀 데이터 저장.
        each_cell.value = ut_cell.get_column_letter(each_cell.column)
        # print(ut_cell.get_column_letter(each_cell.column))

        # 각 셀의 정력 값 지정
        each_cell.alignment = Alignment(horizontal='right', vertical='center')  # 수평, 수직 정렬

        # 각 셀의 폰트 값 지정
        # 폰트 지정 : bold = 폰트 두께, name = 폰트 종류, size = 폰트 크기, underline = 밑줄, color = 폰트 색깔
        each_cell.font = Font(bold=True, name='Arial', size=12, underline='single', color='1bb638')

for row in worksheet2.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3) :  # Row 시트에 행 단위로 반복
    print(row)
    for each_cell in row :
        # 각 셀의 행 값을 셀 데이터로 저장.
        each_cell.value = each_cell.row
        # print(each_cell.row)

        # 각 셀의 정렬 값 지정
        each_cell.alignment = Alignment(horizontal='center', vertical='center')  # 수평, 수직 정렬

        # 각 셀의 폰트 값 지정
        # 폰트 지정 : italic = 이텔릭체, name = 폰트 종류, size = 폰트 크기, color = 폰트 색깔
        each_cell.font = Font(italic=True, name='Consoras', size=10, color='ff0000')


# 워크북의 변경 내용을 새로운 파일에 저장
workbook.save('./output/create_workbook2.xlsx')

