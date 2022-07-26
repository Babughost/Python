import openpyxl
import openpyxl.utils.cell as utils_cell # 셀 관련 함수 사용을 위해 import

# 2. 셀(cell) 데이터 다루기
# 셀(cell) 객체 이용해서 엑셀의 셀에 저장된 데이터 변경
# cell 관련 함수 사용하기

file = "./input/df.xlsx" # Excel 파일 경로 지정

workbook = openpyxl.load_workbook(file) # 워크북(Workbook) 객체 생성
worksheet = workbook['Sheet1'] # 시트(sheet) 객체 생성

# 엑셀의 셀(cell)은 행과 열의 위치 값을 가짐.
# 엑셀의 경우 행은 숫자 (1, 2, 3, ... ), 열은 문자 (A, B, C, ... ) 로 표현함.
# openpyxl의 경우 셀의 좌표 값은 [문자숫자] 형태로 사용. 즉 [열행] 순서로 접근.

# 1. 시트객체에서 좌표 값으로 셀 접근
b2 = worksheet['B3']
# 셀 객체는 coordinate : 좌표, column : 열, row : 행, value : 값 속성을 가지고 있음
print(b2, b2.coordinate, b2.column, b2.row, b2.value) # <Cell 'Sheet1'.B2> b2 2 3 2

# 좌표 값에서 열(column)은 문자로 되어 있어서,
# 1) 열을 숫자로 변환 2) 숫자를 열에 해당하는 문자로 변환하는 기능을 제공.
# get_column_letter() : 숫자를 열 문자로 변환
print(utils_cell.get_column_letter(b2.column)) # B
# column_index_from_string() : 열 문자를 숫자로 변환
print(utils_cell.column_index_from_string('B')) # 2
print(utils_cell.column_index_from_string('Z')) # 26
print(utils_cell.column_index_from_string('AA'))    # 27


# 2. 시트객체의 cell() 메서드 사용

c4 = worksheet.cell(row=4, column=3)
print(c4, c4.coordinate, c4.value) # <Cell 'Sheet1'.C4> C4 6

# 셀 데이터 변경

b2.value = 50       # 18번 줄에서 B3를 불러옴
c4.value = 50       # 34번 줄에서 4번째 줄 3번째 행을 불러옴
print(b2.value)
print(c4.value)

# 워크북의 변경내용을 새로운 파일에 저장
workbook.save('./output/df2.xlsx')
