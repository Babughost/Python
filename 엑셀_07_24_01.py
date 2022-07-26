# openpyxl는 외부 모듈이라서 설치 후 사용 가능
# openpyxl 모듈을 사용하려면 package를 설치 해야 함.
# pip install openpyxl

import openpyxl

# 엑셀 사용 1 : openpyxl 이용해서 엑셀 파일 열기

file = "./input/df.xlsx"

# Workbook 객체, Worksheet 객체 사용

# 1. 엑셀 파일을 읽어서 Workbook 객체 생성
# load_workbook() 메서드 사용
workbook = openpyxl.load_workbook(file)
print(workbook)
print(type(workbook))

# 2. 시트 읽기 : 시트 (Sheet) 객체 생성하기.
# Workbook 객체를 사용.
print(workbook.sheetnames) # ['Sheet1']. sheetnames : 해당 워크북의 모든 sheet 반환

# 1) 시트 이름을 사용해서 직접 선택
worksheet = workbook['Sheet1'] # Workbook 객체에서 이름을 key 값으로 시트(Sheet) 객체 생성.
print(worksheet) # <Worksheet "Sheet1">
print(worksheet.title)  # Sheet1. 시트이름

# 2) 활성화된 Sheet 객체 생성하기. 모든 엑셀에는 단 하나의 활성화된 시크가 존재.
active_sheet = workbook.active # 활성화된 sheet를 가져와서 시트(Sheet) 객체 생성.
print(active_sheet) # <Worksheet "Sheet1">. 시트 이름을 사용하는 방법과 객체 동일.